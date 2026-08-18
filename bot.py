import asyncio
import io
import logging
import re
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo

from openpyxl import Workbook

from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup,
    KeyboardButton, ReplyKeyboardMarkup, ReplyKeyboardRemove,
    BotCommand, BotCommandScopeDefault,
)
from telegram.constants import ParseMode
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, ContextTypes, filters,
)

import config
import db

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("botont")

HEX16 = re.compile(r"^[0-9A-Fa-f]{16}$")
TIKET_INSERA = re.compile(r"^INC\d{8}$", re.I)
TIKET_DSC = re.compile(r"^1-[0-9A-Z]{6,10}$", re.I)


# ============================================================
# util
# ============================================================

def kb(rows):
    return InlineKeyboardMarkup(rows)


async def guard(update: Update):
    """Kembalikan record teknisi, atau None kalau tidak terdaftar."""
    uid = update.effective_user.id
    t = await db.get_teknisi(uid)
    if not t or not t["aktif"]:
        return None
    return t


# Enam langkah yang dilihat teknisi. Nomor langkah membuat orang tahu
# posisinya, bukan sekadar nama status internal.
LANGKAH = {
    "ASSIGNED":   (1, "Telepon pelanggan dulu. Pastikan beliau ada di rumah dan "
                      "bersedia ONT-nya diganti."),
    "KENDALA":    (1, "Coba hubungi pelanggan lagi."),
    "CARING_OK":  (2, "Kirim permintaan tiket di grup TSEL. Formatnya sudah "
                      "disiapkan bot, tinggal lengkapi nama dan nomor HP pelanggan."),
    "REQ_TIKET":  (3, "Tunggu admin grup menerbitkan tiket. Kalau nomornya sudah "
                      "keluar, catat di sini."),
    "TIKET_OPEN": (4, "Datangi lokasi, ganti ONT-nya, lalu catat SN perangkat baru "
                      "beserta 2 foto."),
    "GANTI_OK":   (5, "Kirim permintaan config ke helpdesk."),
    "REQ_CONFIG": (5, "Tunggu helpdesk selesai config. Cek internet pelanggan sudah "
                      "menyala."),
    "CONFIG_OK":  (6, "Pastikan internet pelanggan normal, lalu tutup order ini."),
}
TOTAL_LANGKAH = 6


def ringkas_odp(odp) -> str:
    """'ODC-RJW-FDV/22' -> 'RJW-FDV/22'. Kosong -> '-'."""
    if not odp:
        return "-"
    t = str(odp).strip()
    for p in ("ODP-", "ODC-"):
        if t.upper().startswith(p):
            t = t[len(p):]
    return t


def kartu(o, admin: bool = False, sekitar: int = None) -> str:
    lines = [f"<b>No layanan {o['no_inet']}</b>"]

    langkah = LANGKAH.get(o["status"])
    if langkah:
        no, apa = langkah
        lines.append(f"Langkah {no} dari {TOTAL_LANGKAH}")
        lines.append("")
        lines.append(f"<b>Yang harus dilakukan sekarang</b>\n{apa}")
    elif o["status"] == "CLOSED":
        lines.append("Order ini sudah selesai.")

    lines.append("")
    lines.append("<b>Data pelanggan</b>")
    if o["nama_plg"]:
        lines.append(f"Nama      : {o['nama_plg']}")
    if o["cp_plg"]:
        lines.append(f"No HP     : {o['cp_plg']}")
    lines.append(f"ODP       : {ringkas_odp(o['odp'])}")
    lines.append(f"Kecepatan : {o['speed_mb']} Mbps")
    lines.append(f"ONT lama  : {o['type_old'] or '-'}")
    lines.append(f"SN lama   : <code>{o['sn_old'] or '-'}</code>")
    if o["lat"]:
        lines.append(f"Lokasi    : https://maps.google.com/?q={o['lat']},{o['lon']}")

    if o["status"] == "KENDALA":
        lines.append("")
        lines.append("<b>Kendala terakhir</b>")
        lines.append(f"{o['kode_kendala']} — {o['catatan_kendala'] or '-'}")
        if o["followup_date"]:
            lines.append(f"Dijadwalkan ulang: {o['followup_date']:%d %b %Y}")
        if o["percobaan"]:
            lines.append(f"Sudah dicoba {o['percobaan']} kali")

    if o["flag"] == "SUDAH_DUALBAND":
        lines.append("\n⚠ Menurut data, ONT di sini sudah Dual Band. "
                     "Cek dulu di lokasi — kalau benar, pilih kendala K08.")
    elif o["flag"] == "DATA_KOSONG":
        lines.append("\n⚠ Data SN dan tipe ONT lama tidak ada di sistem. "
                     "Catat manual saat di lokasi.")
    elif o["flag"] == "CEK_GEO":
        lines.append("\n⚠ Titik lokasi order ini meragukan. "
                     "Konfirmasi alamat ke pelanggan sebelum berangkat.")
    elif o["flag"] == "MULTI_ONT":
        lines.append("\n⚠ Pelanggan ini punya lebih dari satu ONT.")

    lines.append("")
    lines.append(f"<b>ODP</b> {ringkas_odp(o['odp'])}")
    lines.append(
        f"<b>Wilayah</b> {'Sektor ' + str(o['sektor']) if o['sektor'] else 'luar RJW'}"
        f" · ODC {o['odc'] or '-'}")
    lines.append(f"<b>Kode wilayah</b> <code>{o['group_uid']}</code>")
    if sekitar:
        lines.append(f"Di wilayah yang sama masih ada {sekitar} order Anda yang "
                     "belum selesai — biasanya jaraknya cuma beberapa ratus meter, "
                     "jadi bisa sekalian dikerjakan.")
    elif sekitar == 0:
        lines.append("Ini order terakhir Anda di wilayah ini.")

    if admin:
        lines.append("")
        lines.append("<b>Data sistem</b>")
        lines.append(f"Kode wilayah : <code>{o['group_uid']}</code>")
        lines.append(f"Pemilik : {o['owner_nama'] or 'belum ada'}"
                     + ("  (override)" if o["teknisi_override"] else ""))
        lines.append(f"Status  : {o['status']}")
        lines.append(f"Flag    : {o['flag']}")
    return "\n".join(lines)


async def teks_request_config(no_inet: str) -> str:
    """Format request config yang siap di-forward ke helpdesk."""
    r = await db.pool().fetchrow(
        """SELECT o.no_inet, o.sn_old, o.sn_new, t.no_tiket
           FROM orders o LEFT JOIN tickets t ON t.no_inet = o.no_inet
           WHERE o.no_inet = $1""", no_inet)
    if not r:
        return None
    return ("#request\n"
            f"NO TIKET : {r['no_tiket'] or '-'}\n"
            f"NO LAYANAN : {r['no_inet']}\n"
            f"SN LAMA : {r['sn_old'] or '-'}\n"
            f"SN BARU : {r['sn_new'] or '-'}")


async def teks_struk(no_inet: str) -> str:
    """Struk penyelesaian, siap di-forward atau diarsipkan."""
    r = await db.pool().fetchrow(
        """SELECT o.no_inet, o.sn_old, o.sn_new, o.status, o.type_old,
                  o.nama_plg, o.cp_plg, o.odp, o.jenis_tagih,
                  t.no_tiket, k.nama AS teknisi,
                  to_char(o.closed_at AT TIME ZONE 'Asia/Jakarta',
                          'DD/MM/YYYY HH24:MI') AS selesai
           FROM orders o
           LEFT JOIN tickets t ON t.no_inet = o.no_inet
           LEFT JOIN teknisi k ON k.teknisi_id = o.closed_by
           WHERE o.no_inet = $1""", no_inet)
    if not r:
        return None
    return ("#selesai\n"
            f"NO TIKET : {r['no_tiket'] or '-'}\n"
            f"NO LAYANAN : {r['no_inet']}\n"
            f"SN LAMA : {r['sn_old'] or '-'}\n"
            f"SN BARU : {r['sn_new'] or '-'}\n"
            f"ONT LAMA : {r['type_old'] or '-'}\n"
            f"NAMA PELANGGAN : {r['nama_plg'] or '-'}\n"
            f"CP PELANGGAN : {r['cp_plg'] or '-'}\n"
            f"DATEK ODP : {ringkas_odp(r['odp'])}\n"
            f"TEKNISI : {r['teknisi'] or '-'}\n"
            f"SELESAI : {r['selesai'] or '-'}\n"
            f"TAGIHAN : {r['jenis_tagih'] or '-'}\n"
            f"STATUS : {r['status']}")


def aksi_untuk(status: str, no_inet: str):
    """Tombol yang relevan dengan status saat ini saja."""
    m = {
        "ASSIGNED":   [("✅ Pelanggan setuju, lanjut", "caring"),
                       ("⚠️ Ada kendala", "kendala")],
        "KENDALA":    [("✅ Sekarang bisa dilanjut", "caring"),
                       ("⚠️ Ubah kendala", "kendala")],
        "CARING_OK":  [("✅ Sudah saya kirim di grup TSEL", "reqtiket"),
                       ("⚠️ Ada kendala", "kendala")],
        "REQ_TIKET":  [("📝 Catat nomor tiket", "tiket"),
                       ("⚠️ Ada kendala", "kendala")],
        "TIKET_OPEN": [("📝 Catat SN ONT baru", "sn"),
                       ("⚠️ Ada kendala", "kendala")],
        "GANTI_OK":   [("✅ Sudah kirim ke helpdesk", "reqconfig"),
                       ("🔄 Tampilkan lagi teks request", "fmtconfig")],
        "REQ_CONFIG": [("✅ Internet sudah menyala", "configok")],
        "CONFIG_OK":  [("🏁 Tutup order ini", "close")],
    }
    rows = [[InlineKeyboardButton(t, callback_data=f"{a}|{no_inet}")]
            for t, a in m.get(status, [])]
    rows.append([InlineKeyboardButton("« Kembali ke daftar order", callback_data="list|0")])
    return kb(rows)


# ============================================================
# onboarding
# ============================================================

async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    ok = await db.onboard(u.id, u.username, u.first_name)
    if ok:
        t = await db.get_teknisi(u.id)
        await update.message.reply_text(
            f"Halo {t['nama'].title()}, Anda sudah terdaftar.\n\n"
            "Bot ini menemani Anda mengerjakan penggantian ONT, dari menghubungi "
            "pelanggan sampai order ditutup. Bot yang mengingatkan langkah "
            "berikutnya, jadi Anda tidak perlu menghafal urutannya.\n\n"
            "Ketik /order untuk mulai.\n"
            "Ketik /bantuan kalau bingung.")
    elif await db.is_admin(u.id):
        await update.message.reply_text(
            "Mode admin aktif. Ketik /adminhelp untuk daftar perintah.")
    else:
        await update.message.reply_text(
            "Nomor Telegram Anda belum terdaftar di sistem ini.\n\n"
            f"Kirim nomor berikut ke Officer RJW supaya didaftarkan:\n"
            f"<code>{u.id}</code>\n\n"
            "(tekan nomornya untuk menyalin)",
            parse_mode=ParseMode.HTML)


async def cmd_bantuan(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "<b>Urutan mengerjakan satu order</b>\n"
        "1. Telepon pelanggan, pastikan setuju diganti\n"
        "2. Minta tiket di grup TSEL (formatnya disiapkan bot)\n"
        "3. Catat nomor tiket di bot\n"
        "4. Ganti ONT, catat SN baru + 2 foto\n"
        "5. Minta config ke helpdesk\n"
        "6. Tutup order\n\n"
        "Setiap selesai satu langkah, tekan tombolnya. Bot akan langsung "
        "memberitahu langkah berikutnya.\n\n"
        "<b>Perintah sehari-hari</b>\n"
        "/order — lihat order yang harus dikerjakan hari ini\n"
        "/sisa — berapa order Anda yang belum selesai\n"
        "/cari 1234567890 — buka order tertentu dengan nomor layanannya\n"
        "/klaster RJW-FCB-01 — lihat semua order dalam satu wilayah\n\n"
        "<b>Kalau order Anda habis</b>\n"
        "/ambil — kirim lokasi, bot tunjukkan pekerjaan terdekat yang belum "
        "ada yang pegang\n"
        "/klaimsaya — daftar wilayah yang sedang Anda pegang\n\n"
        "<b>Kalau salah ketik</b>\n"
        "/batal — membatalkan pengisian yang sedang berjalan\n"
        "/struk 1234567890 — cetak ulang bukti order yang sudah selesai\n\n"
        "Kalau ada yang tidak jelas atau bot terasa aneh, hubungi Officer RJW.",
        parse_mode=ParseMode.HTML)


async def cmd_batal(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ada = ctx.user_data.pop("pending", None)
    await update.message.reply_text(
        "Pengisian dibatalkan. Ketik /order untuk kembali ke daftar."
        if ada else
        "Tidak ada pengisian yang sedang berjalan. Ketik /order untuk mulai.")


# ============================================================
# daftar order
# ============================================================

async def kirim_daftar(target, uid: int, ctx):
    kuota = int(await db.get_setting("kuota_harian", "3"))
    rows = await db.antrian(uid, kuota)
    total = await db.pool().fetchval(
        """SELECT COUNT(*) FROM v_order_owner
           WHERE teknisi_id=$1 AND status NOT IN ('CLOSED','BATAL')""", uid)
    if not rows:
        if total:
            return await target.reply_text(
                f"Tidak ada yang bisa dikerjakan hari ini.\n\n"
                f"Anda masih punya {total} order, tapi semuanya sedang menunggu — "
                "entah menunggu tiket terbit, menunggu helpdesk, atau sudah "
                "dijadwalkan ulang ke hari lain.\n\n"
                "Kalau ingin menambah pekerjaan, ketik /ambil.")
        return await target.reply_text(
            "Semua order Anda sudah selesai. Kerja bagus.\n\n"
            "Ketik /ambil untuk mengambil wilayah baru yang belum ada yang pegang.")
    btn = []
    for r in rows:
        no = LANGKAH.get(r["status"], (0, ""))[0]
        btn.append([InlineKeyboardButton(
            f"{ringkas_odp(r['odp'])} · {r['no_inet']} · L{no}",
            callback_data=f"open|{r['no_inet']}")])
    await target.reply_text(
        f"Ini {len(rows)} order untuk hari ini. Sisa keseluruhan {total} order.\n\n"
        "Urutannya sudah dikelompokkan per ODP, jadi yang berdekatan muncul "
        "bersebelahan. L1–L6 menunjukkan sudah sampai langkah berapa.",
        reply_markup=kb(btn))


async def cmd_order(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    t = await guard(update)
    if not t:
        return await update.message.reply_text("Nomor Telegram Anda belum terdaftar. Ketik /start untuk melihat "
            "nomor yang perlu dikirim ke Officer RJW.")
    await kirim_daftar(update.message, t["teknisi_id"], ctx)


async def cmd_sisa(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    t = await guard(update)
    if not t:
        return await update.message.reply_text(
            "Perintah ini untuk teknisi. Sebagai admin, pakai /beban.")
    r = await db.pool().fetchrow(
        "SELECT sisa, kendala, tunggu_tiket, closed FROM v_beban WHERE teknisi_id=$1",
        t["teknisi_id"])
    await update.message.reply_text(
        f"Belum selesai   : {r['sisa']} order\n"
        f"Sudah selesai   : {r['closed']} order\n"
        f"Sedang kendala  : {r['kendala']} order\n"
        f"Menunggu tiket  : {r['tunggu_tiket']} order")


async def cmd_cari(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    t = await guard(update)
    admin = await db.is_admin(uid)
    if not t and not admin:
        return await update.message.reply_text("Nomor Telegram Anda belum terdaftar. Ketik /start untuk melihat "
            "nomor yang perlu dikirim ke Officer RJW.")
    if not ctx.args:
        return await update.message.reply_text("Format: /cari <no_inet>")
    o = await db.get_order(ctx.args[0].strip())
    if not o:
        return await update.message.reply_text("Nomor layanan itu tidak ada di daftar penggantian ONT. "
            "Cek lagi angkanya.")
    if not admin and o["owner_id"] != t["teknisi_id"]:
        return await update.message.reply_text(
            f"Order ini dipegang oleh {o['owner_nama'] or 'teknisi lain'}, "
            "bukan Anda.")
    sekitar = (await db.sisa_sekitar(o["group_uid"], t["teknisi_id"], o["no_inet"])
               if t else None)
    await update.message.reply_text(kartu(o, admin, sekitar), parse_mode=ParseMode.HTML,
                                    reply_markup=aksi_untuk(o["status"], o["no_inet"]),
                                    disable_web_page_preview=True)


async def cmd_struk(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    t = await guard(update)
    admin = await db.is_admin(uid)
    if not t and not admin:
        return await update.message.reply_text("Nomor Telegram Anda belum terdaftar. Ketik /start untuk melihat "
            "nomor yang perlu dikirim ke Officer RJW.")
    if not ctx.args:
        return await update.message.reply_text("Format: /struk <no_inet>")
    no_inet = ctx.args[0].strip()
    o = await db.get_order(no_inet)
    if not o:
        return await update.message.reply_text("Nomor layanan itu tidak ada di daftar penggantian ONT. "
            "Cek lagi angkanya.")
    if not admin and o["owner_id"] != t["teknisi_id"] and o["closed_by"] != t["teknisi_id"]:
        return await update.message.reply_text("Order ini dipegang teknisi lain, jadi tidak bisa Anda kerjakan.")
    if o["status"] != "CLOSED":
        return await update.message.reply_text(
            f"Order ini belum selesai (status: {config.LABEL.get(o['status'], o['status'])}).")
    await update.message.reply_text(await teks_struk(no_inet))


async def cmd_klaster(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    t = await guard(update)
    admin = await db.is_admin(uid)
    if not t and not admin:
        return await update.message.reply_text(
            "Nomor Telegram Anda belum terdaftar. Ketik /start.")
    if not ctx.args:
        return await update.message.reply_text(
            "Format: /klaster <kode wilayah>\n\n"
            "Contoh:\n"
            "/klaster RJW-FCB-01   (satu wilayah)\n"
            "/klaster RJW-FCB      (cari semua wilayah di ODC itu)")

    q = ctx.args[0].strip().upper()
    info = await db.info_klaster(q)
    if not info:
        cocok = await db.cari_klaster(q)
        if not cocok:
            return await update.message.reply_text(
                f"Tidak ada wilayah yang cocok dengan '{q}'.")
        if len(cocok) == 1:
            info = await db.info_klaster(cocok[0]["group_uid"])
        else:
            return await update.message.reply_text(
                f"Ada {len(cocok)} wilayah yang cocok. Pilih salah satu:",
                reply_markup=kb([[InlineKeyboardButton(
                    c["group_uid"], callback_data=f"kls|{c['group_uid']}")]
                    for c in cocok]))
    await kirim_klaster(update.message, info, uid, admin)


async def kirim_klaster(target, info, uid: int, admin: bool):
    rows = await db.isi_klaster(info["group_uid"])
    if not rows:
        return await target.reply_text("Wilayah ini tidak punya order.")

    milik_saya = any(r["teknisi"] and r["teknisi"] for r in rows)
    pemilik = info["pemilik"] or "belum ada yang pegang"
    sisa = sum(1 for r in rows if r["status"] not in ("CLOSED", "BATAL"))
    selesai = sum(1 for r in rows if r["status"] == "CLOSED")

    out = [f"<b>{info['group_uid']}</b>",
           f"Sektor {info['sektor'] or '-'} · ODC {info['odc'] or 'luar RJW'}",
           f"Dipegang: {pemilik}",
           f"Selesai {selesai} dari {len(rows)} order, sisa {sisa}"]
    if info["prioritas"]:
        out.append("★ Didahulukan oleh Officer")
    if admin and info["diam_hari"] is not None:
        out.append(f"Tidak ada aktivitas {info['diam_hari']} hari")
    out.append("")

    odp_now = None
    for r in rows:
        o = ringkas_odp(r["odp"])
        if o != odp_now:
            out.append(f"<b>{o}</b>")
            odp_now = o
        if r["status"] == "CLOSED":
            tag = "selesai"
        elif r["status"] == "KENDALA":
            tag = f"kendala {r['kode_kendala'] or ''}".strip()
            if r["followup_date"]:
                tag += f", ulangi {r['followup_date']:%d %b}"
        else:
            tag = f"langkah {LANGKAH.get(r['status'], (0, ''))[0]}/{TOTAL_LANGKAH}"
        out.append(f"  {r['no_inet']} · {tag}")

    if len(rows) > 30:
        out = out[:80] + ["", f"... dan {len(rows) - 30} order lainnya"]
    await target.reply_text("\n".join(out), parse_mode=ParseMode.HTML)


async def cmd_idgrup(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    c = update.effective_chat
    if c.type == "private":
        return await update.message.reply_text(
            "Perintah ini dipakai di dalam grup, bukan di chat pribadi.\n\n"
            "Buat grup arsip, masukkan bot ini, jadikan admin, lalu ketik "
            "/idgrup di grup tersebut.")
    if not await db.is_admin(update.effective_user.id):
        return
    await update.message.reply_text(
        f"ID grup ini: <code>{c.id}</code>\n\n"
        "Salin, lalu di chat pribadi dengan bot ketik:\n"
        f"<code>/setarsip {c.id}</code>",
        parse_mode=ParseMode.HTML)


async def cmd_setarsip(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    # pemeriksaan admin dilakukan di sini, bukan lewat dekorator, karena
    # fungsi ini didefinisikan sebelum admin_only tersedia
    if not await db.is_admin(update.effective_user.id):
        return
    if not ctx.args:
        kini = await db.get_setting("grup_arsip", "")
        return await update.message.reply_text(
            f"Grup arsip saat ini: {kini or 'belum diatur'}\n\n"
            "Cara mengatur:\n"
            "1. Buat grup Telegram baru\n"
            "2. Masukkan bot ini ke grup, jadikan admin\n"
            "3. Ketik /idgrup di grup tersebut\n"
            "4. Kembali ke sini, ketik /setarsip <id yang didapat>\n\n"
            "Matikan dengan: /setarsip off")
    v = ctx.args[0].strip()
    if v.lower() in ("off", "0", "mati", "-"):
        await db.set_setting("grup_arsip", "", update.effective_user.id)
        return await update.message.reply_text("Pengiriman ke grup arsip dimatikan.")
    try:
        chat_id = int(v)
    except ValueError:
        return await update.message.reply_text(
            "ID grup harus berupa angka, biasanya diawali tanda minus. "
            "Ambil dengan /idgrup di dalam grupnya.")
    try:
        await ctx.bot.send_message(
            chat_id, "Grup ini disetel sebagai arsip order ONT yang sudah selesai.")
    except Exception as e:
        return await update.message.reply_text(
            f"Bot tidak bisa mengirim ke grup itu: {e}\n\n"
            "Pastikan bot sudah dimasukkan ke grup dan dijadikan admin.")
    await db.set_setting("grup_arsip", str(chat_id), update.effective_user.id)
    await update.message.reply_text(
        "Tersimpan. Mulai sekarang tiap order yang ditutup akan dikirim ke grup "
        "itu beserta kedua fotonya.")


async def kirim_arsip(ctx, no_inet: str):
    """Kirim struk + 2 foto ke grup arsip. Kegagalan tidak boleh
    mengganggu penutupan order, jadi semua error ditelan dan dicatat."""
    try:
        gid = await db.get_setting("grup_arsip", "")
        if not gid:
            return
        o = await db.pool().fetchrow(
            """SELECT foto_label_sn, foto_terpasang FROM orders WHERE no_inet=$1""",
            no_inet)
        teks = await teks_struk(no_inet)
        foto = [f for f in (o["foto_label_sn"], o["foto_terpasang"]) if f]
        if len(foto) == 2:
            from telegram import InputMediaPhoto
            await ctx.bot.send_media_group(
                int(gid),
                [InputMediaPhoto(foto[0], caption=teks),
                 InputMediaPhoto(foto[1])])
        elif foto:
            await ctx.bot.send_photo(int(gid), foto[0], caption=teks)
        else:
            await ctx.bot.send_message(int(gid), teks)
    except Exception as e:
        log.warning("gagal kirim arsip %s: %s", no_inet, e)


# ============================================================
# klaim mandiri
# ============================================================

async def batas_klaim(uid: int):
    """None kalau boleh klaim, atau pesan penolakan."""
    maks = int(await db.get_setting("max_klaim_aktif", "2"))
    maks_pegang = int(await db.get_setting("max_klaster_pegang", "4"))
    aktif, pegang = await db.klaim_aktif(uid)
    if pegang >= maks_pegang:
        return (f"Anda memegang {pegang} klaster hasil klaim (batas {maks_pegang}), "
                f"termasuk yang sedang terparkir karena kendala.\n"
                "Selesaikan atau lepas salah satu lewat /klaimsaya.")
    if aktif >= maks:
        return (f"Anda punya {aktif} klaster yang masih bisa dikerjakan "
                f"(batas {maks}). Selesaikan dulu sebelum ambil yang baru.")
    return None


async def cmd_ambil(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    t = await guard(update)
    if not t:
        return await update.message.reply_text("Nomor Telegram Anda belum terdaftar. Ketik /start untuk melihat "
            "nomor yang perlu dikirim ke Officer RJW.")
    tolak = await batas_klaim(t["teknisi_id"])
    if tolak:
        return await update.message.reply_text(tolak)
    await update.message.reply_text(
        "Kirim lokasi Anda sekarang. Tekan tombol di bawah — jangan pilih titik "
        "manual di peta, karena lokasi manual akan ditandai di catatan.",
        reply_markup=ReplyKeyboardMarkup(
            [[KeyboardButton("Kirim lokasi saya", request_location=True)]],
            resize_keyboard=True, one_time_keyboard=True))
    ctx.user_data["pending"] = ("lokasi_klaim",)


async def on_location(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    p = ctx.user_data.get("pending")
    if not p or p[0] != "lokasi_klaim":
        return
    ctx.user_data.pop("pending")
    loc = update.message.location
    live = getattr(loc, "live_period", None) is not None
    ctx.user_data["lokasi"] = (loc.latitude, loc.longitude, live)

    radius = float(await db.get_setting("radius_klaim_km", "3"))
    rows = await db.kolam_terdekat(loc.latitude, loc.longitude, radius)
    if not rows:
        return await update.message.reply_text(
            f"Tidak ada klaster bebas dalam radius {radius:.0f} km dari lokasi Anda.\n"
            "Coba lagi dari lokasi lain, atau kerjakan order yang sudah ada lewat /order.",
            reply_markup=ReplyKeyboardRemove())

    btn = []
    for r in rows:
        tanda = "★ " if r["prioritas"] else ""
        umur = f" · diam {r['diam_hari']}h" if r["diam_hari"] >= 3 else ""
        btn.append([InlineKeyboardButton(
            f"{tanda}{r['group_uid']} · {r['km']:.1f} km · {r['sisa']} order{umur}",
            callback_data=f"klaim|{r['group_uid']}")])
    await update.message.reply_text(
        "Klaster bebas terdekat. Mengambil satu klaster berarti mengambil "
        "seluruh order di dalamnya.\n★ = didorong oleh Officer.",
        reply_markup=ReplyKeyboardRemove())
    await update.message.reply_text("Pilih klaster:", reply_markup=kb(btn))


async def do_klaim(q, ctx, group_uid: str, uid: int):
    tolak = await batas_klaim(uid)
    if tolak:
        return await q.message.reply_text(tolak)
    lat, lon, live = ctx.user_data.get("lokasi", (None, None, None))
    jarak = None
    if lat is not None:
        k = await db.pool().fetchrow(
            "SELECT lat, lon FROM klaster WHERE group_uid=$1", group_uid)
        if k:
            from math import radians, sin, cos, asin, sqrt
            dlat = radians(k["lat"] - lat)
            dlon = radians(k["lon"] - lon)
            jarak = 6371 * 2 * asin(sqrt(
                sin(dlat / 2) ** 2 +
                cos(radians(lat)) * cos(radians(k["lat"])) * sin(dlon / 2) ** 2))
    ok = await db.klaim(group_uid, uid, lat=lat, lon=lon, jarak=jarak, live=live)
    if not ok:
        return await q.message.reply_text(
            f"Klaster {group_uid} baru saja diambil teknisi lain. Coba /ambil lagi.")
    hari = int(await db.get_setting("klaim_expire_hari", "5"))
    await q.message.reply_text(
        f"Klaster {group_uid} sekarang milik Anda.\n"
        f"Tenggat {hari} hari — kalau tidak ada progres sampai batas itu, "
        "klaster otomatis kembali ke kolam.",
        reply_markup=kb([[InlineKeyboardButton("Lihat order", callback_data="list|0")]]))


async def cmd_klaimsaya(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    t = await guard(update)
    if not t:
        return await update.message.reply_text(
            "Perintah ini untuk teknisi. Sebagai admin, pakai /kolam.")
    rows = await db.klaim_saya(t["teknisi_id"])
    if not rows:
        return await update.message.reply_text("Anda belum memegang klaster apa pun.")
    out, btn = ["<b>Klaster yang Anda pegang</b>"], []
    for r in rows:
        asal = "klaim" if r["claim_mode"] == "self" else "penugasan"
        tenggat = f" · tenggat {r['expires_at']:%d %b}" if r["expires_at"] else ""
        out.append(f"{r['group_uid']} · sisa {r['sisa']} · {asal}{tenggat}")
        if r["claim_mode"] == "self":
            btn.append([InlineKeyboardButton(f"Lepas {r['group_uid']}",
                                             callback_data=f"lepas|{r['group_uid']}")])
    await update.message.reply_text("\n".join(out), parse_mode=ParseMode.HTML,
                                    reply_markup=kb(btn) if btn else None)


# ============================================================
# aksi per order
# ============================================================

async def lanjut_caring(msg, no_inet: str, uid: int):
    """Catat caring selesai lalu tampilkan teks request tiket yang sudah lengkap."""
    o = await db.get_order(no_inet)
    await db.transisi(no_inet, "CARING_OK", uid,
                      extra_sql=(", caring_at=now(), kode_kendala=NULL, "
                                 "catatan_kendala=NULL, followup_date=NULL"))
    link = await db.get_setting("link_grup_tsel", "")
    teks = ("Bagus. <b>Langkah 2 dari 6</b>\n\n"
            "Salin teks di bawah ini dan kirim ke grup TSEL. Semua datanya "
            "sudah lengkap, tidak perlu diubah.")
    if link and link != "https://t.me/":
        teks += f"\n\nGrup: {link}"
    await msg.reply_text(teks, parse_mode=ParseMode.HTML)
    await msg.reply_text(
        f"No Internet : {o['no_inet']}\n"
        f"Nama Pelanggan : {o['nama_plg']}\n"
        f"CP Pelanggan : {o['cp_plg']}\n"
        "Witel : BANDUNG\n"
        "STO : RJW\n"
        "Request : Tiket Symptom Z_PERMINTAAN_044 untuk penggantian ONT")
    await msg.reply_text(
        "Kalau sudah terkirim di grup, tekan tombol di bawah.",
        reply_markup=kb([
            [InlineKeyboardButton("✅ Sudah saya kirim di grup TSEL",
                                  callback_data=f"reqtiket|{no_inet}")],
            [InlineKeyboardButton("✏️ Nama/CP salah, perbaiki",
                                  callback_data=f"ubahplg|{no_inet}")]]))


async def on_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    aksi, _, arg = q.data.partition("|")
    uid = q.from_user.id

    if aksi == "list":
        return await kirim_daftar(q.message, uid, ctx)

    if aksi == "kls":
        info = await db.info_klaster(arg)
        if not info:
            return await q.message.reply_text("Wilayah tidak ditemukan.")
        return await kirim_klaster(q.message, info, uid, await db.is_admin(uid))

    if aksi == "klaim":
        return await do_klaim(q, ctx, arg, uid)

    if aksi == "lepas":
        pemilik = await db.pool().fetchval(
            "SELECT teknisi_id FROM assignment WHERE group_uid=$1 AND aktif", arg)
        if pemilik != uid and not await db.is_admin(uid):
            return await q.message.reply_text("Wilayah ini bukan Anda yang pegang.")
        jalan = await db.pool().fetchval(
            """SELECT COUNT(*) FROM orders WHERE group_uid=$1
               AND status NOT IN ('NEW','ASSIGNED','CLOSED','BATAL')""", arg)
        if jalan:
            return await q.message.reply_text(
                f"Tidak bisa dilepas: {jalan} order di klaster ini sudah jalan "
                "(caring atau lebih). Selesaikan dulu, atau minta Officer melepasnya.")
        await db.lepas(arg, uid)
        return await q.message.reply_text(f"Klaster {arg} dikembalikan ke kolam.")

    if aksi == "open":
        o = await db.get_order(arg)
        if not o:
            return await q.message.reply_text("Nomor layanan itu tidak ada di daftar penggantian ONT. "
            "Cek lagi angkanya.")
        sekitar = (await db.sisa_sekitar(o["group_uid"], uid, o["no_inet"])
                   if o["owner_id"] == uid else None)
        return await q.message.reply_text(
            kartu(o, await db.is_admin(uid), sekitar), parse_mode=ParseMode.HTML,
            reply_markup=aksi_untuk(o["status"], o["no_inet"]),
            disable_web_page_preview=True)

    # kd dan tagih membawa argumen kedua: no_inet|nilai
    kode_kendala = jenis_tagih = None
    if aksi == "kd":
        arg, _, kode_kendala = arg.partition("|")
    elif aksi == "tagih":
        arg, _, jenis_tagih = arg.partition("|")
        if jenis_tagih not in ("MANHOUR", "REKON"):
            return await q.message.reply_text("Pilihan tidak dikenali.")

    o = await db.get_order(arg)
    if not o:
        return await q.message.reply_text("Nomor layanan itu tidak ada di daftar penggantian ONT. "
            "Cek lagi angkanya.")
    if o["owner_id"] != uid and not await db.is_admin(uid):
        return await q.message.reply_text("Order ini dipegang teknisi lain, jadi tidak bisa Anda kerjakan.")

    # ---- caring OK ----
    if aksi == "caring":
        if not (o["nama_plg"] and o["cp_plg"]):
            ctx.user_data["pending"] = ("nama_plg", arg)
            return await q.message.reply_text(
                "Sebelum minta tiket, bot perlu nama pelanggannya.\n\n"
                "Ketik nama pelanggan sesuai data pelanggan yang Anda buka. "
                "Nanti bot yang merangkai teks requestnya, jadi Anda tinggal "
                "salin dan kirim ke grup.\n\nBatal? Ketik /batal.")
        return await lanjut_caring(q.message, arg, uid)

    if aksi == "ubahplg":
        ctx.user_data["pending"] = ("nama_plg", arg)
        return await q.message.reply_text("Ketik ulang nama pelanggannya.")

    # ---- tandai sudah request tiket ----
    if aksi == "reqtiket":
        await db.transisi(arg, "REQ_TIKET", uid,
                          extra_sql=", req_tiket_at=now(), req_tiket_by=$3",
                          extra_args=(uid,))
        sla = await db.get_setting("sla_tiket_jam", "8")
        return await q.message.reply_text(
            f"Tercatat. <b>Langkah 3 dari 6</b>\n\n"
            "Sekarang tunggu admin grup menerbitkan tiketnya. Begitu nomor tiket "
            "muncul di grup, kembali ke sini dan catat nomornya.\n\n"
            f"Kalau lewat {sla} jam belum keluar juga, bot akan mengingatkan Anda. "
            "Sambil menunggu, Anda bisa mengerjakan order lain lewat /order.",
            parse_mode=ParseMode.HTML,
            reply_markup=kb([[InlineKeyboardButton("📝 Catat nomor tiket",
                                                   callback_data=f"tiket|{arg}")]]))

    # ---- minta input teks ----
    if aksi == "tiket":
        ctx.user_data["pending"] = ("tiket", arg)
        return await q.message.reply_text(
            "Ketik nomor tiket yang diterbitkan admin grup.\n\n"
            "Contoh yang benar:\n"
            "INC51892061  (tiket INSERA)\n"
            "1-S9Y14GE  (tiket DSC)\n\n"
            "Salin persis seperti yang tertulis di grup. Batal? Ketik /batal.")

    if aksi == "sn":
        ctx.user_data["pending"] = ("sn", arg)
        return await q.message.reply_text(
            "Ketik SN ONT <b>yang baru dipasang</b>.\n\n"
            "SN ada di stiker belakang atau bawah perangkat, 16 karakter, "
            "biasanya diawali huruf. Jangan ketik SN yang lama.\n\n"
            f"SN lama order ini: <code>{o['sn_old'] or 'tidak ada di data'}</code>\n\n"
            "Batal? Ketik /batal.",
            parse_mode=ParseMode.HTML)

    if aksi == "kendala":
        rows = await db.pool().fetch("SELECT kode,label FROM kendala_ref ORDER BY urutan")
        return await q.message.reply_text(
            "Apa yang menghambat pekerjaan ini? Pilih yang paling sesuai:",
            reply_markup=kb([[InlineKeyboardButton(f"{r['kode']} {r['label']}",
                                                   callback_data=f"kd|{arg}|{r['kode']}")]
                             for r in rows]))

    if aksi == "kd":
        ref = await db.pool().fetchrow("SELECT * FROM kendala_ref WHERE kode=$1", kode_kendala)
        if ref is None:
            return await q.message.reply_text("Kode kendala tidak dikenali.")
        if ref["perlu_catatan"] or ref["perlu_tanggal"]:
            ctx.user_data["pending"] = ("kendala", arg, kode_kendala, ref["perlu_tanggal"])
            if ref["perlu_tanggal"]:
                besok = (date.today() + timedelta(days=1)).isoformat()
                minta = (f"Kapan pelanggan minta didatangi lagi?\n"
                         f"Ketik tanggalnya dengan format tahun-bulan-tanggal, "
                         f"contoh: {besok}")
            else:
                minta = ("Tulis keterangan singkat supaya Officer paham "
                         "situasinya. Contoh: rumah dikunci, tetangga bilang "
                         "pindah bulan lalu.")
            return await q.message.reply_text(f"{ref['label']}.\n\n{minta}")
        return await simpan_kendala(q.message, arg, kode_kendala, uid, None, None)

    if aksi == "fmtconfig":
        teks = await teks_request_config(arg)
        await q.message.reply_text(teks)
        return await q.message.reply_text(
            "Teruskan pesan di atas ke helpdesk, lalu tekan tombol di bawah.",
            reply_markup=kb([[InlineKeyboardButton("✅ Sudah kirim ke helpdesk",
                                                   callback_data=f"reqconfig|{arg}")]]))

    if aksi == "reqconfig":
        await db.transisi(arg, "REQ_CONFIG", uid, extra_sql=", req_config_at=now()")
        return await q.message.reply_text(
            "Tercatat. Sekarang tunggu helpdesk menyelesaikan config.\n\n"
            "Cek internet pelanggan — kalau sudah bisa browsing, tekan tombol di bawah.",
            reply_markup=kb([[InlineKeyboardButton("✅ Internet sudah menyala",
                                                   callback_data=f"configok|{arg}")]]))

    if aksi == "configok":
        await db.transisi(arg, "CONFIG_OK", uid, extra_sql=", config_at=now()")
        return await q.message.reply_text(
            "Bagus. <b>Langkah 6 dari 6</b>\n\n"
            "Sebelum menutup, pastikan sekali lagi internet pelanggan benar-benar "
            "normal dan beliau tidak mengeluhkan apa pun.",
            parse_mode=ParseMode.HTML,
            reply_markup=kb([[InlineKeyboardButton("🏁 Tutup order ini",
                                                   callback_data=f"close|{arg}")]]))

    if aksi == "close":
        if not o["sn_new"]:
            return await q.message.reply_text(
                "Order ini belum bisa ditutup karena SN ONT baru belum dicatat.",
                reply_markup=kb([[InlineKeyboardButton("📝 Catat SN ONT baru",
                                                       callback_data=f"sn|{arg}")]]))
        if not (o["foto_label_sn"] and o["foto_terpasang"]):
            kurang = ("foto stiker SN" if not o["foto_label_sn"]
                      else "foto perangkat terpasang")
            return await q.message.reply_text(
                f"Order ini belum bisa ditutup karena {kurang} belum ada.\n\n"
                "Kirim fotonya sekarang.")
        return await q.message.reply_text(
            "Terakhir: pekerjaan ini ditagihkan sebagai apa?",
            reply_markup=kb([
                [InlineKeyboardButton("💰 Manhour", callback_data=f"tagih|{arg}|MANHOUR")],
                [InlineKeyboardButton("📋 Rekon", callback_data=f"tagih|{arg}|REKON")]]))

    if aksi == "tagih":
        await db.transisi(arg, "CLOSED", uid, catatan=f"tagih {jenis_tagih}",
                          extra_sql=", closed_at=now(), closed_by=$3, jenis_tagih=$4",
                          extra_args=(uid, jenis_tagih))
        await db.pool().execute(
            "UPDATE tickets SET closed_at=now() WHERE no_inet=$1", arg)
        await q.message.reply_text(await teks_struk(arg))
        await kirim_arsip(ctx, arg)
        return await q.message.reply_text(
            "Order selesai. Terima kasih.\n\n"
            "Simpan bukti di atas — kalau nanti ada yang menanyakan, "
            "bisa dipanggil lagi dengan /struk.",
            reply_markup=kb([[InlineKeyboardButton("Lanjut order berikutnya",
                                                   callback_data="list|0")]]))


async def simpan_kendala(msg, no_inet, kode, uid, catatan, tgl):
    await db.transisi(
        no_inet, "KENDALA", uid, kode_kendala=kode, catatan=catatan,
        extra_sql=(", kode_kendala=$3, catatan_kendala=$4, followup_date=$5, "
                   "percobaan=percobaan+1"),
        extra_args=(kode, catatan, tgl))
    lanjut = (f"Order ini akan muncul lagi di daftar Anda pada {tgl:%d %b %Y}."
              if tgl else
              "Order ini akan muncul lagi di daftar Anda besok.")
    await msg.reply_text(
        f"Kendala tercatat. {lanjut}\n\n"
        "Officer bisa melihat catatan ini, jadi Anda tidak perlu melapor terpisah.",
        reply_markup=kb([[InlineKeyboardButton("« Kembali ke daftar order",
                                               callback_data="list|0")]]))


# ============================================================
# input teks & foto
# ============================================================

async def on_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    p = ctx.user_data.get("pending")
    if not p:
        return
    uid = update.effective_user.id
    teks = update.message.text.strip()
    jenis = p[0]

    # ---- nama pelanggan ----
    if jenis == "nama_plg":
        no_inet = p[1]
        if len(teks) < 3:
            return await update.message.reply_text(
                "Nama terlalu pendek. Ketik nama pelanggan selengkapnya, "
                "atau /batal.")
        await db.pool().execute(
            "UPDATE orders SET nama_plg=$2, updated_at=now() WHERE no_inet=$1",
            no_inet, teks.upper())
        ctx.user_data["pending"] = ("cp_plg", no_inet)
        return await update.message.reply_text(
            "Sekarang ketik nomor HP pelanggannya.\n"
            "Contoh: 081380874793")

    # ---- CP pelanggan ----
    if jenis == "cp_plg":
        no_inet = p[1]
        cp = re.sub(r"[^0-9+]", "", teks)
        if len(re.sub(r"\D", "", cp)) < 9:
            return await update.message.reply_text(
                f"\"{teks}\" tidak terlihat seperti nomor HP. "
                "Ketik ulang angkanya saja, atau /batal.")
        await db.pool().execute(
            "UPDATE orders SET cp_plg=$2, updated_at=now() WHERE no_inet=$1",
            no_inet, cp)
        ctx.user_data.pop("pending")
        return await lanjut_caring(update.message, no_inet, uid)

    # ---- nomor tiket ----
    if jenis == "tiket":
        no_inet = p[1]
        val = teks.upper().replace(" ", "")
        if TIKET_INSERA.match(val):
            tipe = "INSERA"
        elif TIKET_DSC.match(val):
            tipe = "DSC"
        else:
            return await update.message.reply_text(
                f"Nomor \"{teks}\" belum sesuai format tiket.\n\n"
                "Tiket INSERA: huruf INC diikuti 8 angka, contoh INC51892061\n"
                "Tiket DSC: angka 1, tanda hubung, lalu 6-10 huruf/angka, "
                "contoh 1-S9Y14GE\n\n"
                "Coba salin ulang dari grup, atau ketik /batal.")
        dupe = await db.pool().fetchval(
            "SELECT no_inet FROM tickets WHERE no_tiket=$1 AND no_inet<>$2", val, no_inet)
        if dupe:
            return await update.message.reply_text(
                f"Nomor tiket ini sudah tercatat untuk order lain ({dupe}).\n\n"
                "Kemungkinan tersalin dari pekerjaan sebelumnya. Cek lagi nomor "
                "tiket untuk order ini di grup.")
        o = await db.get_order(no_inet)
        await db.pool().execute(
            """INSERT INTO tickets(no_inet,no_tiket,jenis,requested_by,requested_at)
               VALUES($1,$2,$3,$4,$5)
               ON CONFLICT (no_inet) DO UPDATE SET
                 no_tiket=EXCLUDED.no_tiket, jenis=EXCLUDED.jenis, issued_at=now()""",
            no_inet, val, tipe, uid, o["req_tiket_at"])
        await db.transisi(no_inet, "TIKET_OPEN", uid, catatan=f"{tipe} {val}",
                          extra_sql=", tiket_at=now()")
        ctx.user_data.pop("pending")
        return await update.message.reply_text(
            f"Tiket {tipe} {val} tercatat. <b>Langkah 4 dari 6</b>\n\n"
            "Silakan datangi lokasi dan ganti ONT-nya. Setelah terpasang dan "
            "internet menyala, kembali ke sini untuk mencatat SN perangkat baru "
            "dan mengirim 2 foto.",
            parse_mode=ParseMode.HTML,
            reply_markup=kb([[InlineKeyboardButton("📝 Catat SN ONT baru",
                                                   callback_data=f"sn|{no_inet}")]]))

    # ---- SN baru ----
    if jenis == "sn":
        no_inet = p[1]
        sn = teks.upper().replace(" ", "").replace(":", "")
        o = await db.get_order(no_inet)
        if not HEX16.match(sn):
            return await update.message.reply_text(
                f"SN yang Anda ketik ada {len(sn)} karakter, seharusnya tepat 16 "
                "dan hanya berisi angka 0-9 dan huruf A sampai F.\n\n"
                "Cek lagi stiker di perangkat. Hati-hati membedakan angka 0 dengan "
                "huruf O, dan angka 1 dengan huruf I.\n\n"
                "Ketik ulang, atau /batal.")
        vendor = next((v for v, p8 in config.SN_PREFIX.items() if sn.startswith(p8)), None)
        if not vendor:
            return await update.message.reply_text(
                "SN ini tidak dikenali sebagai perangkat HUAWEI, ZTE, maupun "
                "FIBERHOME.\n\n"
                "Biasanya karena ada karakter yang salah baca. Coba foto stikernya "
                "lalu perbesar, dan ketik ulang.\n\n"
                "Kalau perangkatnya memang merek lain, laporkan ke Officer RJW.")
        if o["sn_old"] and sn == o["sn_old"].upper():
            return await update.message.reply_text(
                "SN ini sama persis dengan ONT yang lama.\n\n"
                "Yang perlu dicatat adalah SN perangkat <b>baru</b> yang barusan "
                "Anda pasang, bukan yang dicabut.",
                parse_mode=ParseMode.HTML)
        dipakai = await db.sn_dipakai(sn, no_inet)
        if dipakai:
            return await update.message.reply_text(
                f"SN ini sudah tercatat terpasang di pelanggan lain ({dipakai}).\n\n"
                "Satu perangkat tidak bisa terpasang di dua tempat. Kemungkinan "
                "tersalin dari pekerjaan sebelumnya — cek ulang stiker perangkat "
                "yang baru saja Anda pasang.")
        await db.transisi(no_inet, "GANTI_OK", uid, sn_new=sn,
                          extra_sql=", sn_new=$3, vendor_new=$4, ganti_at=now()",
                          extra_args=(sn, vendor))
        ctx.user_data["pending"] = ("foto_label", no_inet)
        return await update.message.reply_text(
            f"SN {sn} ({vendor}) tercatat.\n\n"
            "Sekarang kirim <b>foto pertama</b>: stiker SN pada perangkat baru. "
            "Pastikan tulisannya terbaca jelas.",
            parse_mode=ParseMode.HTML)

    # ---- kendala dengan keterangan / tanggal ----
    if jenis == "kendala":
        _, no_inet, kode, perlu_tgl = p
        tgl = None
        catatan = teks
        if perlu_tgl:
            try:
                tgl = datetime.strptime(teks, "%Y-%m-%d").date()
            except ValueError:
                besok = (date.today() + timedelta(days=1)).isoformat()
                return await update.message.reply_text(
                    f"Tanggalnya belum sesuai format. Tulis tahun-bulan-tanggal "
                    f"dengan tanda hubung, contoh: {besok}")
            if tgl <= date.today():
                return await update.message.reply_text(
                    "Tanggalnya sudah lewat atau hari ini. Isi tanggal setelah "
                    "hari ini, karena order ini akan dimunculkan lagi pada tanggal itu.")
            catatan = None
        ctx.user_data.pop("pending")
        return await simpan_kendala(update.message, no_inet, kode, uid, catatan, tgl)


async def on_photo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    p = ctx.user_data.get("pending")
    if not p or p[0] not in ("foto_label", "foto_terpasang"):
        return
    fid = update.message.photo[-1].file_id
    no_inet = p[1]
    if p[0] == "foto_label":
        await db.pool().execute(
            "UPDATE orders SET foto_label_sn=$2, updated_at=now() WHERE no_inet=$1",
            no_inet, fid)
        ctx.user_data["pending"] = ("foto_terpasang", no_inet)
        return await update.message.reply_text(
            "Foto pertama tersimpan.\n\n"
            "Sekarang <b>foto kedua</b>: perangkat baru yang sudah terpasang di "
            "lokasi pelanggan, lampunya menyala.",
            parse_mode=ParseMode.HTML)
    await db.pool().execute(
        "UPDATE orders SET foto_terpasang=$2, updated_at=now() WHERE no_inet=$1",
        no_inet, fid)
    ctx.user_data.pop("pending")
    teks = await teks_request_config(no_inet)
    await update.message.reply_text(teks)
    await update.message.reply_text(
        "Foto lengkap. <b>Langkah 5 dari 6</b>\n\n"
        "Teruskan pesan di atas ke helpdesk untuk minta config. Kalau sudah "
        "terkirim, tekan tombol di bawah.",
        parse_mode=ParseMode.HTML,
        reply_markup=kb([[InlineKeyboardButton("✅ Sudah kirim ke helpdesk",
                                               callback_data=f"reqconfig|{no_inet}")]]))


# ============================================================
# admin
# ============================================================

def admin_only(fn):
    async def wrap(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
        if not await db.is_admin(update.effective_user.id):
            return
        return await fn(update, ctx)
    return wrap


BANTUAN_ADMIN = {
 "progres": ("Sejauh mana keseluruhan program, dan kapan kira-kira selesai.\n\n"
   "Menampilkan selesai vs total, laju rata-rata 7 hari terakhir, lalu "
   "memperkirakan sisa hari kerja dari laju itu. Angka perkiraan inilah yang "
   "bisa dipakai untuk melapor ke atasan — bukan perkiraan, tapi hitungan "
   "dari data nyata.\n\nPakai seminggu sekali."),
 "hariini": ("Apa yang terjadi hari ini, dibandingkan kemarin.\n\n"
   "Yang penting dibaca bukan angka close saja, tapi perbandingan antar tahap. "
   "Kalau caring 80 tapi close 3 selama beberapa hari, berarti ada sumbatan di "
   "tengah dan itu bukan salah teknisi.\n\nPakai tiap sore."),
 "tunggutiket": ("Order yang sudah minta tiket tapi tiketnya belum terbit, "
   "diurut dari yang paling lama menunggu.\n\n"
   "Kalau daftar ini panjang dan umurnya puluhan jam, hambatannya ada di admin "
   "grup TSEL, bukan di teknisi. Ini bukti angka kalau nanti target meleset."),
 "beban": ("Sisa order per teknisi: sisa, kendala, tunggu tiket.\n\n"
   "Yang dicari bukan angka sisa — awalnya semua sekitar 80. Yang dicari nama "
   "dengan angka kendala jauh di atas rata-rata. Tanda ⚠ berarti orang itu "
   "belum menekan /start, jadi bot belum bisa mengirim apa pun kepadanya."),
 "teknisi": ("Melihat pekerjaan satu orang.\n\n"
   "Format: /teknisi <nama|nik>\n"
   "Contoh: /teknisi CECEP\n\n"
   "Menampilkan wilayah yang dipegangnya, berapa selesai, berapa kendala, dan "
   "apakah wilayah itu hasil penugasan awal atau diambil sendiri lewat /ambil.\n\n"
   "Tambahkan kata 'detail' untuk melihat daftar nomor layanannya:\n"
   "/teknisi CECEP detail"),
 "stagnan": ("Klaster yang tidak ada progres sama sekali selama 7 hari.\n\n"
   "Pengaman terhadap teknisi yang berhenti bekerja tanpa memberi tahu — cuti, "
   "sakit, resign. Tanpa ini, 80 order bisa diam sebulan tanpa ketahuan."),
 "export": ("Seluruh data ke Excel, 4 sheet.\n\n"
   "Untuk pertanyaan 'mana saja'. Sheet DETAIL berisi semua order dengan filter "
   "otomatis — untuk melihat yang kendala, filter kolom status = KENDALA."),
 "rekap": ("Jumlah order per status, plus urutan kendala terbanyak.\n\n"
   "Pandangan sekilas. Untuk detail pakai /export."),
 "assign": ("Memindahkan satu klaster (sekitar 10 order) ke teknisi lain.\n\n"
   "Contoh: /assign RPLONT-04D26E BUDI\n\n"
   "Untuk penyeimbangan kecil. Bot memperingatkan kalau ada order yang sudah "
   "melewati tahap minta tiket — tiket itu tetap tercatat atas nama perequest "
   "aslinya, karena di grup TSEL tiket terikat ke orang."),
 "pindahorder": ("Memindahkan SATU order saja, tanpa mengganggu wilayahnya.\n\n"
   "Format: /pindahorder <no_layanan> <nama|nik>\n"
   "Contoh: /pindahorder 131177179142 DIKI SODIKIN\n"
   "Batalkan: /pindahorder 131177179142 batal\n\n"
   "Pakai ini untuk pengecualian — pelanggan minta teknisi tertentu, atau akses "
   "lokasi yang butuh orang khusus. Untuk penyeimbangan beban, /assign per "
   "wilayah lebih baik karena menjaga kedekatan lokasi.\n\n"
   "Lihat semua yang sedang dikunci dengan /dikunci."),
 "dikunci": ("Daftar order yang dipindah satuan lewat /pindahorder.\n\n"
   "Perlu diperiksa berkala — order yang dikunci tidak ikut berpindah kalau "
   "wilayahnya nanti dialihkan, jadi bisa tertinggal tanpa ada yang sadar."),
 "pindahzona": ("Memindahkan seluruh zona, sekitar 8 klaster atau 80 order.\n\n"
   "Contoh: /pindahzona Z14 BUDI\n\nUntuk teknisi resign atau mutasi."),
 "nonaktif": ("Menghentikan distribusi pagi ke seorang teknisi.\n\n"
   "Contoh: /nonaktif BUDI\n\n"
   "Bot akan memberi tahu berapa order yang masih menggantung atas namanya. "
   "Order itu TIDAK pindah otomatis — Anda tetap harus memindahkannya."),
 "setkuota": ("Berapa order dikirim ke tiap teknisi tiap pagi. Sekarang bisa "
   "dicek dengan mengetik /setkuota tanpa angka.\n\n"
   "Terlalu kecil, teknisi kehabisan pekerjaan. Terlalu besar, daftarnya "
   "tenggelam di chat dan malah tidak dikerjakan."),
 "onboarding": ("Siapa saja yang belum menekan /start.\n\n"
   "Selama nama masih ada di daftar ini, bot tidak bisa mengirim pesan apa pun "
   "ke orang tersebut — Telegram melarang bot memulai percakapan duluan."),
 "kolam": ("Klaster yang tidak ada pemiliknya, siap diambil teknisi lewat "
   "/ambil.\n\n"
   "Sekarang isinya nol karena semua klaster sudah terbagi ke 65 teknisi. "
   "Kolam terisi kalau ada klaim kedaluwarsa, atau Anda sengaja melepas."),
 "lepaspaksa": ("Mencabut klaster dari pemiliknya, kembalikan ke kolam.\n\n"
   "Contoh: /lepaspaksa RPLONT-04D26E\n\n"
   "Dipakai kalau teknisi menghilang dan Anda belum tahu harus dialihkan ke "
   "siapa — biarkan siapa pun yang lewat mengambilnya."),
 "tambahteknisi": ("Mendaftarkan teknisi baru.\n\n"
   "Format: /tambahteknisi <nomor_telegram> <nik> <nama lengkap>\n"
   "Contoh: /tambahteknisi 473776150 18980067 AHMAD RIZAL\n\n"
   "Nomor Telegram-nya didapat dari orangnya sendiri: suruh dia kirim /start "
   "ke bot ini, bot akan membalas dengan nomornya. Setelah terdaftar, dia "
   "belum punya pekerjaan — beri lewat /assign atau /pindahzona."),
 "setarsip": ("Mengatur grup Telegram penampung bukti order selesai.\n\n"
   "Tiap order yang ditutup dikirim ke grup itu berisi struk dan dua fotonya, "
   "jadi Anda bisa memeriksa hasil kerja tanpa menghubungi teknisi.\n\n"
   "Caranya: buat grup, masukkan bot, jadikan admin, ketik /idgrup di grup itu, "
   "lalu /setarsip <id> di sini. Matikan dengan /setarsip off.\n\n"
   "Kalau pengiriman gagal, order tetap tertutup normal — kegagalan arsip "
   "tidak pernah menghambat teknisi."),
 "setsektor": ("Mengubah sektor seorang teknisi.\n\n"
   "Format: /setsektor <nama|nik> <1|2|3|bebas>\n"
   "Contoh: /setsektor AHMAD RIZAL 1\n\n"
   "Teknisi sektor hanya melihat wilayah sektornya di /ambil, sampai sisa "
   "sektornya turun di bawah 50 order. Teknisi bebas melihat semua sektor, "
   "dengan sektor paling tertinggal muncul lebih dulu."),
 "tambahadmin": ("Memberi akses admin ke seseorang.\n\n"
   "Format: /tambahadmin <nomor_telegram> <nama>\n\n"
   "Admin bisa memakai semua perintah pemantauan dan penugasan, termasuk "
   "memindahkan pekerjaan orang. Beri hanya ke yang memang mengelola program."),
 "hapusadmin": ("Mencabut akses admin.\n\n"
   "Nomornya dilihat di /daftaradmin. Anda tidak bisa mencabut akses diri "
   "sendiri. Admin yang berasal dari setelan Railway (BOOTSTRAP_ADMINS) juga "
   "tidak bisa dicabut dari sini — harus lewat Railway."),
 "aktifkan": ("Mengaktifkan kembali teknisi yang pernah dinonaktifkan.\n\n"
   "Format: /aktifkan <nama|nik>\n\n"
   "Order yang dulu atas namanya tidak otomatis kembali — kalau sudah "
   "dipindahkan, harus dipindahkan lagi secara manual."),
 "dorong": ("Memaksa satu klaster muncul di puncak daftar /ambil semua "
   "teknisi, dengan tanda bintang, mengabaikan batas jarak 3 km.\n\n"
   "Contoh: /dorong RPLONT-04D26E\nBatalkan: /dorong RPLONT-04D26E off\n\n"
   "Alat untuk wilayah pinggiran yang tidak ada yang mau."),
}


@admin_only
async def cmd_adminhelp(update: Update, ctx):
    if ctx.args:
        k = ctx.args[0].lstrip("/").lower()
        if k in BANTUAN_ADMIN:
            return await update.message.reply_text(
                f"<b>/{k}</b>\n\n{BANTUAN_ADMIN[k]}", parse_mode=ParseMode.HTML)
        return await update.message.reply_text(
            f"Tidak ada penjelasan untuk '{k}'. Ketik /adminhelp untuk daftarnya.")
    await update.message.reply_text(
        "Ketik <code>/adminhelp namaperintah</code> untuk penjelasan lengkap "
        "satu perintah. Contoh: <code>/adminhelp stagnan</code>\n\n"
        "<b>Tiap hari</b>\n"
        "/hariini — apa yang terjadi hari ini vs kemarin\n"
        "/tunggutiket — order mandek menunggu tiket terbit\n\n"
        "<b>Tiap minggu</b>\n"
        "/progres — sejauh mana, dan perkiraan kapan selesai\n"
        "/beban — sebaran beban dan siapa yang banyak kendala\n"
        "/teknisi &lt;nama&gt; — wilayah dan order satu teknisi\n"
        "/stagnan — klaster yang tidak bergerak seminggu\n"
        "/export — semua data ke Excel untuk ditelusuri\n"
        "/rekap — ringkasan status dan kendala\n\n"
        "<b>Saat ada kejadian</b>\n"
        "/assign — pindahkan 1 wilayah ke teknisi lain\n"
        "/pindahorder — pindahkan 1 order saja\n"
        "/dikunci — daftar order yang dipindah satuan\n"
        "/pindahzona — pindahkan 1 zona penuh\n"
        "/nonaktif — hentikan distribusi ke seorang teknisi\n"
        "/setkuota — atur jatah order harian\n"
        "/onboarding — siapa yang belum menekan /start\n\n"
        "<b>Kelola pengguna</b>\n"
        "/tambahteknisi — daftarkan teknisi baru\n"
        "/setsektor — ubah sektor teknisi\n"
        "/setarsip — grup penampung bukti order selesai\n"
        "/aktifkan — aktifkan kembali teknisi nonaktif\n"
        "/tambahadmin — beri akses admin\n"
        "/hapusadmin — cabut akses admin\n"
        "/daftaradmin — siapa saja yang punya akses admin\n\n"
        "<b>Kolam pekerjaan bebas</b>\n"
        "/sektor — sisa order per sektor\n"
        "/kolam — klaster tanpa pemilik\n"
        "/lepaspaksa — cabut klaster, kembalikan ke kolam\n"
        "/dorong — paksa 1 klaster ke puncak daftar semua teknisi\n\n"
        "<b>Satuan</b>\n"
        "/cari 1234567890 — buka satu order\n"
        "/struk 1234567890 — cetak ulang bukti order selesai",
        parse_mode=ParseMode.HTML)


@admin_only
async def cmd_beban(update: Update, ctx):
    rows = await db.beban()
    out = ["<b>Sisa order per teknisi</b>", "<code>sisa kdl tiket  nama</code>"]
    for r in rows[:70]:
        mark = "" if r["siap_dm"] else " ⚠"
        out.append(f"<code>{r['sisa']:>4} {r['kendala']:>3} {r['tunggu_tiket']:>5}</code>  {r['nama']}{mark}")
    out.append("\n⚠ = belum tekan /start, bot belum bisa kirim DM")
    await update.message.reply_text("\n".join(out), parse_mode=ParseMode.HTML)


@admin_only
async def cmd_assign(update: Update, ctx):
    if len(ctx.args) < 2:
        return await update.message.reply_text("Format: /assign <group_uid> <nama|nik>")
    gu, q = ctx.args[0].strip().upper(), " ".join(ctx.args[1:])
    if not await db.klaster_ada(gu):
        mirip = await db.cari_klaster(gu.replace("/", "-").rsplit("-", 1)[0])
        pesan = f"Tidak ada wilayah berkode {gu}."
        if "/" in ctx.args[0]:
            pesan += ("\n\nYang Anda ketik itu kode ODP — pakai garis miring. "
                      "Kode wilayah memakai tanda hubung, contoh RJW-FDX-10.")
        if mirip:
            pesan += "\n\nMungkin yang Anda maksud:\n" + "\n".join(
                f"<code>{m['group_uid']}</code>" for m in mirip[:8])
        pesan += "\n\nCek isinya dulu dengan /klaster sebelum memindahkan."
        return await update.message.reply_text(pesan, parse_mode=ParseMode.HTML)
    cand = await db.cari_teknisi(q)
    if len(cand) != 1:
        pesan = ("Teknisi tidak ditemukan." if not cand
                 else "Nama tidak unik: " + ", ".join(c["nama"] for c in cand))
        return await update.message.reply_text(pesan)
    blok = await db.pool().fetch(
        """SELECT no_inet,status FROM orders WHERE group_uid=$1
           AND status NOT IN ('NEW','ASSIGNED','CARING_OK','KENDALA','CLOSED','BATAL')""", gu)
    lama = await db.pool().fetchval(
        """SELECT t.nama FROM assignment a JOIN teknisi t ON t.teknisi_id=a.teknisi_id
           WHERE a.group_uid=$1 AND a.aktif""", gu)
    n = await db.pool().fetchval(
        """SELECT COUNT(*) FROM orders WHERE group_uid=$1
           AND status NOT IN ('CLOSED','BATAL')""", gu)
    await db.set_assignment(gu, cand[0]["teknisi_id"], update.effective_user.id)
    msg = (f"Wilayah {gu} ({n} order belum selesai) "
           f"{'dipindah dari ' + lama + ' ' if lama else ''}→ {cand[0]['nama']}.")
    if blok:
        msg += (f"\n\n⚠ {len(blok)} order sudah melewati tahap request tiket dan tetap "
                "tercatat atas nama perequest asli: "
                + ", ".join(f"{b['no_inet']}({b['status']})" for b in blok[:5]))
    await update.message.reply_text(msg)


@admin_only
async def cmd_pindahzona(update: Update, ctx):
    if len(ctx.args) < 2:
        return await update.message.reply_text("Format: /pindahzona <zona> <nama|nik>")
    zona, q = ctx.args[0].upper(), " ".join(ctx.args[1:])
    cand = await db.cari_teknisi(q)
    if len(cand) != 1:
        return await update.message.reply_text("Teknisi tidak unik/ketemu.")
    kl = await db.klaster_zona(zona)
    for k in kl:
        await db.set_assignment(k["group_uid"], cand[0]["teknisi_id"],
                                update.effective_user.id, catatan=f"pindah zona {zona}")
    await update.message.reply_text(
        f"{len(kl)} klaster di zona {zona} → {cand[0]['nama']}.")


@admin_only
async def cmd_stagnan(update: Update, ctx):
    hari = int(await db.get_setting("stagnan_hari", "7"))
    rows = await db.stagnan(hari)
    if not rows:
        return await update.message.reply_text(f"Tidak ada klaster diam ≥ {hari} hari.")
    out = [f"<b>Klaster tanpa progres ≥ {hari} hari</b>"]
    for r in rows:
        out.append(f"{r['group_uid']} · {r['nama'] or '-'} · sisa {r['sisa']} · diam {r['diam_hari']}h")
    await update.message.reply_text("\n".join(out), parse_mode=ParseMode.HTML)


@admin_only
async def cmd_rekap(update: Update, ctx):
    f = await db.funnel()
    p = await db.pareto_kendala()
    out = ["<b>Funnel status</b>"]
    for r in f:
        out.append(f"{config.LABEL.get(r['status'], r['status'])}: {r['n']}")
    if p:
        out.append("\n<b>Pareto kendala</b>")
        for r in p:
            out.append(f"{r['kode_kendala']} {r['label']}: {r['n']}")
    await update.message.reply_text("\n".join(out), parse_mode=ParseMode.HTML)


@admin_only
async def cmd_tunggutiket(update: Update, ctx):
    rows = await db.aging("REQ_TIKET", "req_tiket_at")
    if not rows:
        return await update.message.reply_text("Tidak ada order menunggu tiket.")
    out = [f"<b>Menunggu tiket ({len(rows)} terlama)</b>"]
    for r in rows[:25]:
        out.append(f"{r['no_inet']} · {r['nama'] or '-'} · {r['jam']:.0f} jam")
    await update.message.reply_text("\n".join(out), parse_mode=ParseMode.HTML)


@admin_only
async def cmd_onboarding(update: Update, ctx):
    rows = await db.belum_onboarding()
    if not rows:
        return await update.message.reply_text("Semua teknisi aktif sudah onboarding.")
    out = [f"<b>Belum tekan /start ({len(rows)})</b>"]
    for r in rows:
        out.append(f"{r['nik']} · {r['nama']} · <code>{r['teknisi_id']}</code>")
    await update.message.reply_text("\n".join(out), parse_mode=ParseMode.HTML)


@admin_only
async def cmd_perbaiki(update: Update, ctx):
    n = await db.hapus_assignment_hantu()
    await update.message.reply_text(
        f"{n} penugasan yang menunjuk wilayah tidak ada telah dihapus."
        if n else "Tidak ada penugasan bermasalah.")


@admin_only
async def cmd_teknisi(update: Update, ctx):
    if not ctx.args:
        return await update.message.reply_text(
            "Format: /teknisi <nama|nik>\n\n"
            "Contoh: /teknisi CECEP\n\n"
            "Menampilkan wilayah yang dipegang orang itu beserta progresnya. "
            "Tambahkan kata 'detail' di akhir untuk melihat daftar ordernya:\n"
            "/teknisi CECEP detail")
    args = list(ctx.args)
    detail = args and args[-1].lower() in ("detail", "rinci", "order")
    if detail:
        args.pop()
    cand = await db.cari_teknisi(" ".join(args))
    if len(cand) != 1:
        return await update.message.reply_text(
            "Teknisi tidak ditemukan." if not cand
            else "Nama tidak unik: " + ", ".join(c["nama"] for c in cand))
    t = cand[0]
    info = await db.get_teknisi(t["teknisi_id"])
    kl = await db.pegangan_teknisi(t["teknisi_id"])

    peran = f"Sektor {info['sektor']}" if info["sektor"] else "bebas (semua sektor)"
    siap = "sudah" if info["onboarded_at"] else "BELUM tekan /start"
    out = [f"<b>{t['nama']}</b> · {t['nik']}",
           f"Peran: {peran} · {siap}"]
    if not kl:
        out.append("\nBelum memegang wilayah apa pun.")
        return await update.message.reply_text("\n".join(out), parse_mode=ParseMode.HTML)

    tot = sum(k["total"] for k in kl)
    sisa = sum(k["sisa"] for k in kl)
    kdl = sum(k["kendala"] for k in kl)
    out.append(f"{len(kl)} wilayah · {tot - sisa} selesai dari {tot} · "
               f"sisa {sisa} · kendala {kdl}\n")

    for k in kl:
        asal = "klaim sendiri" if k["claim_mode"] == "self" else "penugasan"
        if k["ada_override"] and not k["claim_mode"]:
            asal = "dikunci manual"
        out.append(f"<code>{k['group_uid']}</code> · S{k['sektor'] or '-'} · "
                   f"{asal}\n  {k['selesai']}/{k['total']} selesai · "
                   f"sisa {k['sisa']}"
                   + (f" · kendala {k['kendala']}" if k["kendala"] else ""))

    if detail:
        rows = await db.order_teknisi(t["teknisi_id"])
        out.append(f"\n<b>Order belum selesai ({len(rows)})</b>")
        gu = None
        for r in rows[:60]:
            if r["group_uid"] != gu:
                out.append(f"<code>{r['group_uid']}</code>")
                gu = r["group_uid"]
            tag = (f"kendala {r['kode_kendala']}" if r["status"] == "KENDALA"
                   else f"L{LANGKAH.get(r['status'], (0, ''))[0]}")
            out.append(f"  {ringkas_odp(r['odp'])} · {r['no_inet']} · {tag}")
        if len(rows) > 60:
            out.append(f"... dan {len(rows) - 60} lainnya. Pakai /export untuk semuanya.")
    else:
        out.append("\nTambahkan 'detail' untuk melihat daftar ordernya.")

    teks = "\n".join(out)
    for i in range(0, len(teks), 3800):
        await update.message.reply_text(teks[i:i + 3800], parse_mode=ParseMode.HTML)


@admin_only
async def cmd_pindahorder(update: Update, ctx):
    if len(ctx.args) < 2:
        return await update.message.reply_text(
            "Format: /pindahorder &lt;no_layanan&gt; &lt;nama|nik&gt;\n\n"
            "Contoh:\n/pindahorder 131177179142 DIKI SODIKIN\n\n"
            "Memindahkan <b>satu order saja</b>, wilayahnya tetap dipegang "
            "pemilik lama. Untuk memindahkan seluruh wilayah, pakai /assign.\n\n"
            "Batalkan dengan: /pindahorder &lt;no_layanan&gt; batal",
            parse_mode=ParseMode.HTML)

    no_inet = ctx.args[0].strip()
    o = await db.get_order(no_inet)
    if not o:
        return await update.message.reply_text(
            "Nomor layanan itu tidak ada di daftar penggantian ONT.")

    if ctx.args[1].lower() in ("batal", "off", "hapus", "-"):
        if not o["teknisi_override"]:
            return await update.message.reply_text(
                "Order ini memang tidak dikunci ke siapa pun.")
        await db.set_override(no_inet, None)
        balik = await db.get_order(no_inet)
        return await update.message.reply_text(
            f"Kunci dilepas. Order {no_inet} kembali mengikuti wilayahnya, "
            f"sekarang dipegang {balik['owner_nama'] or 'belum ada siapa-siapa'}.")

    if o["status"] in ("CLOSED", "BATAL"):
        return await update.message.reply_text(
            "Order ini sudah selesai, tidak perlu dipindah.")

    cand = await db.cari_teknisi(" ".join(ctx.args[1:]))
    if len(cand) != 1:
        return await update.message.reply_text(
            "Teknisi tidak ditemukan." if not cand
            else "Nama tidak unik: " + ", ".join(c["nama"] for c in cand))
    t = cand[0]
    if o["owner_id"] == t["teknisi_id"]:
        return await update.message.reply_text(
            f"Order ini memang sudah dipegang {t['nama']}.")

    info = await db.get_teknisi(t["teknisi_id"])
    await db.set_override(no_inet, t["teknisi_id"])
    await db.transisi(no_inet, o["status"] if o["status"] != "NEW" else "ASSIGNED",
                      update.effective_user.id, role="admin",
                      catatan=f"order dipindah ke {t['nama']}",
                      extra_sql=", assigned_at=COALESCE(assigned_at, now())")

    pesan = [f"Order {no_inet} sekarang dipegang {t['nama']}.",
             f"Wilayah {o['group_uid']} tetap milik "
             f"{o['owner_nama'] or 'belum ada siapa-siapa'}."]
    if o["status"] not in ("NEW", "ASSIGNED"):
        pesan.append(f"\n⚠ Order ini sudah jalan sampai "
                     f"{config.LABEL.get(o['status'], o['status'])}. "
                     "Kalau tiketnya sudah terbit, tiket itu tetap tercatat atas "
                     "nama perequest aslinya.")
    if o["sektor"] and info["sektor"] and o["sektor"] != info["sektor"]:
        pesan.append(f"\n⚠ Order ini di Sektor {o['sektor']}, sedangkan "
                     f"{t['nama']} teknisi Sektor {info['sektor']}.")
    if not info["onboarded_at"]:
        pesan.append(f"\n⚠ {t['nama']} belum menekan /start, jadi belum bisa "
                     "menerima pesan dari bot.")
    await update.message.reply_text("\n".join(pesan))


@admin_only
async def cmd_dikunci(update: Update, ctx):
    rows = await db.daftar_override()
    if not rows:
        return await update.message.reply_text(
            "Tidak ada order yang dikunci ke teknisi tertentu.")
    out = [f"<b>Order dipindah satuan ({len(rows)})</b>", ""]
    for r in rows:
        out.append(f"{r['no_inet']} · {r['dikunci_ke']}")
        out.append(f"  wilayah {r['group_uid']} milik {r['nama_klaster'] or '-'}")
    out.append("\nLepas dengan: /pindahorder &lt;no_layanan&gt; batal")
    await update.message.reply_text("\n".join(out), parse_mode=ParseMode.HTML)


@admin_only
async def cmd_kolam(update: Update, ctx):
    r = await db.ringkas_kolam()
    rows = await db.pool().fetch(
        "SELECT group_uid, zona, sisa, diam_hari, prioritas FROM v_kolam "
        "ORDER BY diam_hari DESC, sisa DESC LIMIT 20")
    out = [f"<b>Kolam</b>: {r['klaster']} klaster · {r['order_sisa']} order · "
           f"{r['didorong']} didorong · terlama diam {r['terlama']} hari", ""]
    for k in rows:
        out.append(f"{'★ ' if k['prioritas'] else ''}{k['group_uid']} · {k['zona']} · "
                   f"sisa {k['sisa']} · diam {k['diam_hari']}h")
    await update.message.reply_text("\n".join(out), parse_mode=ParseMode.HTML)


@admin_only
async def cmd_dorong(update: Update, ctx):
    if not ctx.args:
        return await update.message.reply_text(
            "Format: /dorong <group_uid>  ·  batalkan: /dorong <group_uid> off")
    gu = ctx.args[0]
    nyala = not (len(ctx.args) > 1 and ctx.args[1].lower() in ("off", "0", "batal"))
    if not await db.set_prioritas(gu, nyala):
        return await update.message.reply_text("Klaster tidak ditemukan.")
    await update.message.reply_text(
        f"Klaster {gu} {'didorong ke puncak daftar semua teknisi' if nyala else 'tidak lagi didorong'}.")


@admin_only
async def cmd_lepaspaksa(update: Update, ctx):
    if not ctx.args:
        return await update.message.reply_text("Format: /lepaspaksa <group_uid>")
    gu = ctx.args[0]
    pemilik = await db.pool().fetchval(
        "SELECT teknisi_id FROM assignment WHERE group_uid=$1 AND aktif", gu)
    if not pemilik:
        return await update.message.reply_text("Klaster ini sudah tidak ada pemiliknya.")
    await db.lepas(gu, pemilik, aksi="dilepas_admin")
    await update.message.reply_text(f"Klaster {gu} dikembalikan ke kolam.")


@admin_only
async def cmd_hariini(update: Update, ctx):
    hari_ini = await db.produksi_hari(0)
    kemarin = await db.produksi_hari(1)
    kmap = {r["status_to"]: r["n"] for r in kemarin}
    out = ["<b>Aktivitas hari ini</b>", "<code>hari ini  kemarin  tahap</code>"]
    if not hari_ini:
        out.append("Belum ada aktivitas.")
    for r in hari_ini:
        lab = config.LABEL.get(r["status_to"], r["status_to"])
        out.append(f"<code>{r['n']:>8} {kmap.get(r['status_to'], 0):>9}</code>  {lab}")
    top = await db.top_teknisi_hari()
    if top:
        out.append("\n<b>Order ditutup hari ini</b>")
        for t in top:
            out.append(f"{t['n']:>3}  {t['nama']}")
    await update.message.reply_text("\n".join(out), parse_mode=ParseMode.HTML)


@admin_only
async def cmd_progres(update: Update, ctx):
    p = await db.progres_kumulatif()
    laju = await db.laju_harian(7) or 0
    sisa = p["total"] - p["closed"]
    pct = 100 * p["closed"] / p["total"] if p["total"] else 0
    out = [
        "<b>Progres keseluruhan</b>",
        f"Selesai      : {p['closed']} / {p['total']}  ({pct:.1f}%)",
        f"Sedang jalan : {p['jalan']}",
        f"Kendala      : {p['kendala']}",
        f"Belum assign : {p['belum_assign']}",
        "",
        f"Laju 7 hari terakhir: {laju:.1f} order/hari",
    ]
    if laju > 0:
        hk = sisa / laju
        out.append(f"Perkiraan sisa waktu: {hk:.0f} hari kerja "
                   f"(~{hk/26:.1f} bulan) pada laju ini")
    else:
        out.append("Belum bisa memperkirakan sisa waktu — belum ada order ditutup.")
    tren = await db.tren_harian(7)
    if tren:
        out.append("\n<b>7 hari terakhir</b>")
        for t in tren:
            out.append(f"{t['tgl']:%d %b} : {t['n']}")
    await update.message.reply_text("\n".join(out), parse_mode=ParseMode.HTML)


def _bangun_workbook(rows, funnel, pareto, beban, tren, tagih) -> bytes:
    """Blocking. Dipanggil lewat asyncio.to_thread supaya polling tidak berhenti."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DETAIL"
    kolom = list(rows[0].keys())
    ws.append(kolom)
    for r in rows:
        ws.append([r[c] for c in kolom])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("RINGKASAN")
    ws2.append(["Status", "Jumlah"])
    for r in funnel:
        ws2.append([config.LABEL.get(r["status"], r["status"]), r["n"]])
    ws2.append([])
    ws2.append(["Kode kendala", "Keterangan", "Jumlah"])
    for r in pareto:
        ws2.append([r["kode_kendala"], r["label"], r["n"]])

    ws3 = wb.create_sheet("PER_TEKNISI")
    ws3.append(["Teknisi", "Sisa", "Kendala", "Tunggu tiket", "Selesai", "Total"])
    for r in beban:
        ws3.append([r["nama"], r["sisa"], r["kendala"], r["tunggu_tiket"],
                    r["closed"], r["total"]])

    ws4 = wb.create_sheet("TREN_HARIAN")
    ws4.append(["Tanggal", "Order ditutup"])
    for r in tren:
        ws4.append([r["tgl"], r["n"]])

    ws5 = wb.create_sheet("PENAGIHAN")
    ws5.append(["Teknisi", "NIK", "Manhour", "Rekon", "Belum ditandai", "Total selesai"])
    for r in tagih:
        ws5.append([r["teknisi"], r["nik"], r["manhour"], r["rekon"],
                    r["belum_ditandai"], r["total_closed"]])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@admin_only
async def cmd_export(update: Update, ctx):
    await update.message.reply_text("Menyiapkan file, tunggu sebentar...")
    try:
        rows = await db.detail_export()
        if not rows:
            return await update.message.reply_text("Tidak ada data.")
        paket = await asyncio.to_thread(
            _bangun_workbook, rows,
            await db.funnel(), await db.pareto_kendala(),
            await db.beban(), await db.tren_harian(30), await db.rekap_tagih())
        nama = f"MONITORING_ONT_RJW_{datetime.now(ZoneInfo(config.TZ)):%Y%m%d_%H%M}.xlsx"
        await update.message.reply_document(
            document=io.BytesIO(paket), filename=nama,
            caption=f"{len(rows)} order · {len(paket)/1e6:.1f} MB")
    except Exception as e:
        log.exception("export gagal")
        await update.message.reply_text(
            f"Export gagal: {type(e).__name__} — {e}\n"
            "Detail lengkapnya ada di Deploy Logs.")


@admin_only
async def cmd_tambahteknisi(update: Update, ctx):
    if len(ctx.args) < 3 or not ctx.args[0].isdigit():
        return await update.message.reply_text(
            "Format: /tambahteknisi <nomor_telegram> <nik> <nama> [sektor]\n\n"
            "Contoh teknisi sektor 1:\n"
            "/tambahteknisi 473776150 18980067 AHMAD RIZAL 1\n\n"
            "Contoh teknisi bebas (tanpa angka sektor di akhir):\n"
            "/tambahteknisi 473776150 18980067 AHMAD RIZAL\n\n"
            "Nomor Telegram didapat dari orangnya: suruh dia kirim /start ke bot "
            "ini, nanti bot membalas dengan nomornya.")
    uid, nik = int(ctx.args[0]), ctx.args[1]
    sisa = list(ctx.args[2:])
    sektor = None
    if len(sisa) > 1 and sisa[-1] in ("1", "2", "3"):
        sektor = int(sisa.pop())
    nama = " ".join(sisa).upper()
    hasil = await db.tambah_teknisi(uid, nik, nama)
    if hasil == "nik_dipakai":
        return await update.message.reply_text(
            f"NIK {nik} sudah dipakai teknisi lain. Cek lagi NIK-nya.")
    await db.set_sektor(uid, sektor)
    peran = f"teknisi Sektor {sektor}" if sektor else "teknisi bebas (semua sektor)"
    if hasil == "diperbarui":
        return await update.message.reply_text(
            f"Data {nama} diperbarui sebagai {peran} dan diaktifkan kembali.")
    await update.message.reply_text(
        f"{nama} ({nik}) terdaftar sebagai {peran}.\n\n"
        "Dua langkah lagi:\n"
        "1. Suruh dia kirim /start ke bot supaya bisa menerima pesan\n"
        "2. Beri dia pekerjaan — /assign atau /pindahzona, atau biarkan dia "
        "mengambil sendiri lewat /ambil kalau kolam ada isinya")


@admin_only
async def cmd_setsektor(update: Update, ctx):
    if len(ctx.args) < 2:
        return await update.message.reply_text(
            "Format: /setsektor <nama|nik> <1|2|3|bebas>\n\n"
            "Contoh:\n"
            "/setsektor AHMAD RIZAL 1\n"
            "/setsektor 18980067 bebas\n\n"
            "Teknisi sektor hanya bisa mengambil wilayah di sektornya sampai "
            "sektornya hampir habis. Teknisi bebas bisa ke sektor mana pun.")
    nilai = ctx.args[-1].lower()
    if nilai in ("1", "2", "3"):
        sektor = int(nilai)
    elif nilai in ("bebas", "0", "-", "none"):
        sektor = None
    else:
        return await update.message.reply_text(
            "Sektornya harus 1, 2, 3, atau kata 'bebas'.")
    cand = await db.cari_teknisi(" ".join(ctx.args[:-1]))
    if len(cand) != 1:
        return await update.message.reply_text(
            "Teknisi tidak ditemukan." if not cand
            else "Nama tidak unik: " + ", ".join(c["nama"] for c in cand))
    await db.set_sektor(cand[0]["teknisi_id"], sektor)
    await update.message.reply_text(
        f"{cand[0]['nama']} sekarang "
        + (f"teknisi Sektor {sektor}." if sektor
           else "teknisi bebas — bisa mengambil wilayah di sektor mana pun."))


@admin_only
async def cmd_tambahadmin(update: Update, ctx):
    if len(ctx.args) < 2 or not ctx.args[0].isdigit():
        return await update.message.reply_text(
            "Format: /tambahadmin <nomor_telegram> <nama>\n\n"
            "Contoh:\n/tambahadmin 473776150 BUDI SANTOSO\n\n"
            "Admin bisa memakai semua perintah pemantauan dan penugasan. "
            "Beri akses ini hanya ke orang yang memang mengelola program.")
    uid = int(ctx.args[0])
    nama = " ".join(ctx.args[1:]).upper()
    baru = await db.tambah_admin(uid, nama)
    await update.message.reply_text(
        f"{nama} sekarang admin. Suruh dia kirim /start ke bot."
        if baru else f"Data admin {nama} diperbarui.")


@admin_only
async def cmd_hapusadmin(update: Update, ctx):
    if not ctx.args or not ctx.args[0].isdigit():
        return await update.message.reply_text(
            "Format: /hapusadmin <nomor_telegram>\n"
            "Nomornya bisa dilihat di /daftaradmin.")
    uid = int(ctx.args[0])
    if uid == update.effective_user.id:
        return await update.message.reply_text(
            "Tidak bisa menghapus akses admin Anda sendiri.")
    if await db.hapus_admin(uid):
        return await update.message.reply_text("Akses admin dicabut.")
    await update.message.reply_text(
        "Nomor itu tidak ada di daftar admin.\n\n"
        "Kalau dia tetap bisa memakai perintah admin, berarti nomornya "
        "tercantum di variabel BOOTSTRAP_ADMINS di Railway — hapus dari sana.")


@admin_only
async def cmd_daftaradmin(update: Update, ctx):
    rows = await db.daftar_admin()
    out = ["<b>Admin</b>"]
    for r in rows:
        out.append(f"{r['nama']} · <code>{r['teknisi_id']}</code>")
    if not rows:
        out.append("Belum ada admin yang ditambahkan lewat bot.")
    if config.BOOTSTRAP_ADMINS:
        out.append("\n<b>Admin bawaan (dari setelan Railway)</b>")
        for a in config.BOOTSTRAP_ADMINS:
            out.append(f"<code>{a}</code>")
        out.append("\nYang ini hanya bisa dihapus lewat variabel "
                   "BOOTSTRAP_ADMINS di Railway.")
    await update.message.reply_text("\n".join(out), parse_mode=ParseMode.HTML)


@admin_only
async def cmd_aktifkan(update: Update, ctx):
    if not ctx.args:
        return await update.message.reply_text("Format: /aktifkan <nama|nik>")
    cand = await db.pool().fetch(
        """SELECT teknisi_id, nik, nama, aktif FROM teknisi
           WHERE nama ILIKE '%'||$1||'%' OR nik=$1 OR teknisi_id::TEXT=$1
           ORDER BY nama LIMIT 10""", " ".join(ctx.args))
    if len(cand) != 1:
        return await update.message.reply_text(
            "Teknisi tidak ditemukan." if not cand
            else "Nama tidak unik: " + ", ".join(c["nama"] for c in cand))
    if cand[0]["aktif"]:
        return await update.message.reply_text(
            f"{cand[0]['nama']} memang sudah aktif.")
    await db.aktifkan_teknisi(cand[0]["teknisi_id"])
    await update.message.reply_text(
        f"{cand[0]['nama']} diaktifkan kembali dan akan menerima distribusi pagi.")


@admin_only
async def cmd_sektor(update: Update, ctx):
    rows = await db.sisa_sektor()
    out = ["<b>Sisa order per sektor</b>", "<code>sisa   kdl  tek  per org</code>"]
    for r in rows:
        per = r["sisa"] / r["teknisi_sektor"] if r["teknisi_sektor"] else 0
        out.append(f"<code>{r['sisa']:>4} {r['kendala']:>5} {r['teknisi_sektor']:>4}"
                   f" {per:>7.0f}</code>  Sektor {r['sektor']}")
    luar = await db.pool().fetchval(
        """SELECT COUNT(*) FROM orders
           WHERE sektor IS NULL AND status NOT IN ('CLOSED','BATAL')""")
    dibuka = await db.luar_rjw_dibuka()
    out.append(f"\nLuar RJW: {luar} order — "
               + ("sudah dibuka" if dibuka else "belum dibuka, menunggu order RJW habis"))
    out.append("\nKolom 'per org' hanya menghitung teknisi sektor. "
               "37 teknisi bebas menyebar ke sektor mana pun, dan diarahkan "
               "ke sektor dengan sisa terbanyak.")
    await update.message.reply_text("\n".join(out), parse_mode=ParseMode.HTML)


@admin_only
async def cmd_setkuota(update: Update, ctx):
    if not ctx.args or not ctx.args[0].isdigit():
        k = await db.get_setting("kuota_harian", "3")
        return await update.message.reply_text(f"Kuota saat ini: {k}. Ubah: /setkuota <n>")
    await db.set_setting("kuota_harian", ctx.args[0], update.effective_user.id)
    await update.message.reply_text(f"Kuota harian → {ctx.args[0]} order per teknisi.")


@admin_only
async def cmd_nonaktif(update: Update, ctx):
    if not ctx.args:
        return await update.message.reply_text("Format: /nonaktif <nama|nik>")
    cand = await db.cari_teknisi(" ".join(ctx.args))
    if len(cand) != 1:
        return await update.message.reply_text("Teknisi tidak unik/ketemu.")
    await db.pool().execute("UPDATE teknisi SET aktif=FALSE WHERE teknisi_id=$1",
                            cand[0]["teknisi_id"])
    sisa = await db.pool().fetchval(
        """SELECT COUNT(*) FROM v_order_owner
           WHERE teknisi_id=$1 AND status NOT IN ('CLOSED','BATAL')""",
        cand[0]["teknisi_id"])
    await update.message.reply_text(
        f"{cand[0]['nama']} dinonaktifkan. {sisa} order masih atas namanya — "
        "pindahkan dengan /pindahzona atau /assign.")


# ============================================================
# job harian
# ============================================================

async def job_distribusi(ctx: ContextTypes.DEFAULT_TYPE):
    """Push antrian pagi ke tiap teknisi yang sudah onboarding."""
    kuota = int(await db.get_setting("kuota_harian", "3"))
    rows = await db.pool().fetch(
        "SELECT teknisi_id, nama FROM teknisi WHERE aktif AND onboarded_at IS NOT NULL")
    for t in rows:
        ordr = await db.antrian(t["teknisi_id"], kuota)
        if not ordr:
            continue
        btn = [[InlineKeyboardButton(
            f"{ringkas_odp(r['odp'])} · {r['no_inet']} · "
            f"L{LANGKAH.get(r['status'], (0, ''))[0]}",
            callback_data=f"open|{r['no_inet']}")] for r in ordr]
        try:
            await ctx.bot.send_message(
                t["teknisi_id"],
                f"Selamat pagi {t['nama'].title()}.\n\n"
                f"Ini {len(ordr)} order untuk hari ini. Tekan salah satu untuk mulai.",
                reply_markup=kb(btn))
            await db.pool().execute(
                "UPDATE orders SET dikirim_at=now() WHERE no_inet=ANY($1::text[])",
                [r["no_inet"] for r in ordr])
        except Exception as e:
            log.warning("gagal kirim ke %s: %s", t["teknisi_id"], e)


async def job_expire(ctx: ContextTypes.DEFAULT_TYPE):
    """Lepas klaim mandiri yang lewat tenggat dan kembalikan ke kolam."""
    for r in await db.klaim_kedaluwarsa():
        await db.lepas(r["group_uid"], r["teknisi_id"], aksi="kedaluwarsa")
        try:
            await ctx.bot.send_message(
                r["teknisi_id"],
                f"Klaster {r['group_uid']} ({r['sisa']} order tersisa) tidak ada "
                f"aktivitas selama {r['diam_hari']} hari dan dikembalikan ke kolam. "
                "Ambil lagi lewat /ambil kalau masih ingin mengerjakannya.")
        except Exception as e:
            log.warning("notifikasi kedaluwarsa gagal %s: %s", r["teknisi_id"], e)


async def job_sla(ctx: ContextTypes.DEFAULT_TYPE):
    """Ingatkan order yang mandek menunggu tiket atau config."""
    sla_t = int(await db.get_setting("sla_tiket_jam", "8"))
    rows = await db.pool().fetch(
        """SELECT o.no_inet, v.teknisi_id,
                  EXTRACT(EPOCH FROM now()-o.req_tiket_at)/3600 AS jam
           FROM orders o JOIN v_order_owner v ON v.no_inet=o.no_inet
           WHERE o.status='REQ_TIKET' AND o.req_tiket_at < now() - make_interval(hours => $1)""",
        sla_t)
    for r in rows:
        try:
            await ctx.bot.send_message(
                r["teknisi_id"],
                f"Order {r['no_inet']} sudah {r['jam']:.0f} jam menunggu tiket.\n\n"
                "Kalau nomor tiketnya sudah keluar di grup, catat di sini supaya "
                "bisa dilanjutkan. Kalau belum keluar juga, tanyakan ke admin grup.",
                reply_markup=kb([[InlineKeyboardButton(
                    "📝 Catat nomor tiket", callback_data=f"tiket|{r['no_inet']}")]]))
        except Exception as e:
            log.warning("sla dm gagal %s: %s", r["teknisi_id"], e)


# ============================================================

async def on_error(update: object, ctx: ContextTypes.DEFAULT_TYPE):
    """Tanpa ini, error di handler mana pun hanya masuk log dan pengguna
    melihat pesan menggantung tanpa penjelasan."""
    log.exception("handler error", exc_info=ctx.error)
    if not (isinstance(update, Update) and update.effective_message):
        return
    try:
        if await db.is_admin(update.effective_user.id):
            pesan = (f"Kesalahan sistem: {type(ctx.error).__name__} — {ctx.error}\n"
                     "Detail lengkap ada di Deploy Logs.")
        else:
            pesan = ("Maaf, ada gangguan di sistem sehingga perintah tadi tidak "
                     "bisa diproses. Pekerjaan Anda tidak hilang.\n\n"
                     "Coba ulangi sebentar lagi. Kalau masih sama, laporkan ke "
                     "Officer RJW — jangan dikerjakan manual dulu supaya datanya "
                     "tidak dobel.")
        await update.effective_message.reply_text(pesan)
    except Exception:
        pass


MENU_TEKNISI = [
    BotCommand("order",     "Lihat order yang harus dikerjakan hari ini"),
    BotCommand("sisa",      "Berapa order saya yang belum selesai"),
    BotCommand("ambil",     "Ambil wilayah baru kalau order sudah habis"),
    BotCommand("klaimsaya", "Wilayah yang sedang saya pegang"),
    BotCommand("cari",      "Buka order tertentu pakai nomor layanan"),
    BotCommand("klaster",   "Lihat semua order dalam satu wilayah"),
    BotCommand("struk",     "Cetak ulang bukti order yang sudah selesai"),
    BotCommand("batal",     "Batalkan pengisian yang sedang berjalan"),
    BotCommand("bantuan",   "Cara memakai bot ini"),
]


async def post_init(app: Application):
    await db.init()
    log.info("db siap")
    # Menu "/" di Telegram — supaya teknisi tidak perlu menghafal perintah.
    await app.bot.set_my_commands(MENU_TEKNISI, scope=BotCommandScopeDefault())
    log.info("menu perintah dipasang")


async def post_shutdown(app: Application):
    await db.close()


def main():
    app = (Application.builder()
           .token(config.BOT_TOKEN)
           .post_init(post_init)
           .post_shutdown(post_shutdown)
           .build())

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("bantuan", cmd_bantuan))
    app.add_handler(CommandHandler("batal", cmd_batal))
    app.add_handler(CommandHandler("order", cmd_order))
    app.add_handler(CommandHandler("sisa", cmd_sisa))
    app.add_handler(CommandHandler("cari", cmd_cari))
    app.add_handler(CommandHandler("struk", cmd_struk))
    app.add_handler(CommandHandler("klaster", cmd_klaster))
    app.add_handler(CommandHandler("idgrup", cmd_idgrup))
    app.add_handler(CommandHandler("setarsip", cmd_setarsip))
    app.add_handler(CommandHandler("ambil", cmd_ambil))
    app.add_handler(CommandHandler("klaimsaya", cmd_klaimsaya))

    app.add_handler(CommandHandler("adminhelp", cmd_adminhelp))
    app.add_handler(CommandHandler("beban", cmd_beban))
    app.add_handler(CommandHandler("teknisi", cmd_teknisi))
    app.add_handler(CommandHandler("assign", cmd_assign))
    app.add_handler(CommandHandler("pindahorder", cmd_pindahorder))
    app.add_handler(CommandHandler("dikunci", cmd_dikunci))
    app.add_handler(CommandHandler("pindahzona", cmd_pindahzona))
    app.add_handler(CommandHandler("stagnan", cmd_stagnan))
    app.add_handler(CommandHandler("rekap", cmd_rekap))
    app.add_handler(CommandHandler("tunggutiket", cmd_tunggutiket))
    app.add_handler(CommandHandler("onboarding", cmd_onboarding))
    app.add_handler(CommandHandler("tambahteknisi", cmd_tambahteknisi))
    app.add_handler(CommandHandler("aktifkan", cmd_aktifkan))
    app.add_handler(CommandHandler("setsektor", cmd_setsektor))
    app.add_handler(CommandHandler("tambahadmin", cmd_tambahadmin))
    app.add_handler(CommandHandler("hapusadmin", cmd_hapusadmin))
    app.add_handler(CommandHandler("daftaradmin", cmd_daftaradmin))
    app.add_handler(CommandHandler("setkuota", cmd_setkuota))
    app.add_handler(CommandHandler("nonaktif", cmd_nonaktif))
    app.add_handler(CommandHandler("kolam", cmd_kolam))
    app.add_handler(CommandHandler("perbaiki", cmd_perbaiki))
    app.add_handler(CommandHandler("sektor", cmd_sektor))
    app.add_handler(CommandHandler("dorong", cmd_dorong))
    app.add_handler(CommandHandler("lepaspaksa", cmd_lepaspaksa))
    app.add_handler(CommandHandler("hariini", cmd_hariini))
    app.add_handler(CommandHandler("progres", cmd_progres))
    app.add_handler(CommandHandler("export", cmd_export))

    app.add_handler(CallbackQueryHandler(on_callback))
    app.add_handler(MessageHandler(filters.LOCATION, on_location))
    app.add_handler(MessageHandler(filters.PHOTO, on_photo))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))

    app.add_error_handler(on_error)

    jq = app.job_queue
    tz = ZoneInfo(config.TZ)
    jq.run_daily(job_distribusi, time=datetime.strptime("07:30", "%H:%M").time().replace(tzinfo=tz))
    jq.run_repeating(job_sla, interval=timedelta(hours=2), first=timedelta(minutes=5))
    jq.run_daily(job_expire, time=datetime.strptime("06:00", "%H:%M").time().replace(tzinfo=tz))

    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
